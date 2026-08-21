import os
import base64
import json
import time
import datetime
import requests
from email.message import EmailMessage
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from google import genai
from openpyxl import Workbook, load_workbook

load_dotenv()

# Configuration
SPREADSHEET_ID = "1OFy4ZgsUJsY0vwzdbHv-Lq6a6A1fagjb_8dbhX1y5pQ"
SHEET_NAME = "Sheet1"
LOCAL_EXCEL_PATH = "lead_outreach_log.xlsx"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/gmail.compose"
]

SENDER_NAME = "Muhammad Hamza"
SENDER_PHONE = "+92 327 1742800"
SENDER_PORTFOLIO = "https://mrhamza.dev"
SENDER_ROLE = "Full-Stack Web Developer"

ai_client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

def get_google_services():
    creds = None
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server(port=0)
        with open("token.json", "w") as token:
            token.write(creds.to_json())

    sheets_service = build("sheets", "v4", credentials=creds)
    gmail_service = build("gmail", "v1", credentials=creds)
    return sheets_service, gmail_service

def inspect_website(url):
    if not url or not url.startswith("http"):
        return {"exists": False, "notes": "No standalone website"}
    try:
        res = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        soup = BeautifulSoup(res.text, "html.parser")
        title = soup.title.string.strip() if soup.title else "N/A"
        has_meta_viewport = bool(soup.find("meta", attrs={"name": "viewport"}))
        has_form = bool(soup.find("form") or soup.find("iframe"))
        text_content = soup.get_text(separator=" ", strip=True)[:1500]

        return {
            "exists": True,
            "url": url,
            "title": title,
            "mobile_friendly": has_meta_viewport,
            "has_booking_or_contact_form": has_form,
            "snippet": text_content
        }
    except Exception as e:
        return {"exists": False, "error": str(e)}

def analyze_and_draft_email(lead_data, web_audit):
    prompt = f"""
You are an expert web development consultant drafting a high-converting, personalized cold outreach email for Muhammad Hamza ({SENDER_PORTFOLIO}).

Lead Information:
- Business Name: {lead_data['name']}
- Location: {lead_data['location']}
- Industry: {lead_data['industry']}
- Current Website Status: {lead_data['website_status']}
- Social Media: {lead_data['social_link']}
- Web Audit / Research: {json.dumps(web_audit)}

Requirements:
1. Subject line: Short, relevant, and specific to their business/city (no hype or generic sales words).
2. Email Body:
   - Mention their business name and local presence.
   - Point out 2-3 specific, high-impact improvements:
     * If they only have a Facebook/Google listing: Highlight missed Google Search traffic, lack of an automated 24/7 quote/booking form, and improved credibility with a dedicated site.
     * If they have an existing website: Highlight conversion bottlenecks (mobile responsiveness, slow load, booking/contact friction, outdated layout).
   - Keep tone direct, professional, and friendly.
   - State clearly that you build modern web applications and custom booking flows.
   - Include a low-friction call-to-action (e.g., offering a 2-minute visual mockup).
3. Signature block:
{SENDER_NAME}
{SENDER_ROLE}
Phone: {SENDER_PHONE}
Portfolio: {SENDER_PORTFOLIO}

Output Format (strict JSON):
{{
  "subject": "Email Subject",
  "body": "Plain text body"
}}
"""
    response = ai_client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
        config={"response_mime_type": "application/json"}
    )
    return json.loads(response.text)

def create_gmail_draft(gmail_service, to_email, subject, body_text):
    message = EmailMessage()
    message.set_content(body_text)
    message["To"] = to_email
    message["Subject"] = subject

    encoded = base64.urlsafe_b64encode(message.as_bytes()).decode()
    draft = gmail_service.users().drafts().create(
        userId="me",
        body={"message": {"raw": encoded}}
    ).execute()
    return draft.get("id")

def update_local_excel(business_name, email, location, industry, draft_id, subject):
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    headers = ["Business Name", "Email Address", "Location", "Industry", "Draft ID", "Subject Line", "Status", "Timestamp"]
    
    if os.path.exists(LOCAL_EXCEL_PATH):
        try:
            wb = load_workbook(LOCAL_EXCEL_PATH)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Drafted Leads"
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Drafted Leads"
        ws.append(headers)
    
    # Check if entry already exists
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if len(row) > 1 and row[1].value == email:
            row[4].value = draft_id
            row[5].value = subject
            row[6].value = "Draft Created"
            row[7].value = now_str
            found = True
            break
    
    if not found:
        ws.append([business_name, email, location, industry, draft_id, subject, "Draft Created", now_str])
    
    wb.save(LOCAL_EXCEL_PATH)

def run_agent(limit=10):
    sheets_service, gmail_service = get_google_services()

    # Ensure Column I header "Outreach Status" exists in Google Sheet
    try:
        header_res = sheets_service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_ID,
            range=f"{SHEET_NAME}!I1"
        ).execute()
        if not header_res.get("values"):
            sheets_service.spreadsheets().values().update(
                spreadsheetId=SPREADSHEET_ID,
                range=f"{SHEET_NAME}!I1",
                valueInputOption="RAW",
                body={"values": [["Outreach Status"]]}
            ).execute()
    except Exception as e:
        print(f"Warning: Could not check/set header: {e}")

    result = sheets_service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{SHEET_NAME}!A2:I"
    ).execute()
    rows = result.get("values", [])

    processed_count = 0

    for idx, row in enumerate(rows, start=2):
        if processed_count >= limit:
            print(f"Reached batch limit of {limit}. Stopping.")
            break

        business_name = row[0] if len(row) > 0 else ""
        location = row[1] if len(row) > 1 else ""
        industry = row[2] if len(row) > 2 else ""
        phone = row[3] if len(row) > 3 else ""
        email = row[4] if len(row) > 4 else ""
        social = row[5] if len(row) > 5 else ""
        status = row[6] if len(row) > 6 else ""
        outreach_status = row[8] if len(row) > 8 else (row[7] if len(row) > 7 and "Draft" in row[7] else "")

        # Skip entries without valid emails or already processed
        if not email or email == "N/A" or "@" not in email or "Draft Created" in outreach_status or "Draft Created" in status:
            continue

        print(f"\n[{idx}] Processing: {business_name} | {email}")

        # Check website
        website_url = status if "http" in status else None
        audit = inspect_website(website_url)

        lead_info = {
            "name": business_name,
            "location": location,
            "industry": industry,
            "phone": phone,
            "email": email,
            "social_link": social,
            "website_status": status
        }

        # Generate Email
        content = analyze_and_draft_email(lead_info, audit)

        # Create Draft in Gmail
        draft_id = create_gmail_draft(gmail_service, email, content["subject"], content["body"])
        print(f"-> Draft created (ID: {draft_id})")

        # 1. Update Google Sheet Column I (Outreach Status)
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        update_range = f"{SHEET_NAME}!I{idx}"
        sheets_service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=update_range,
            valueInputOption="RAW",
            body={"values": [[f"Draft Created [{now_str}] (ID: {draft_id})"]]}
        ).execute()

        # 2. Update local Excel log file
        update_local_excel(business_name, email, location, industry, draft_id, content["subject"])
        print(f"-> Updated Google Sheet (Column I) & local Excel log ({LOCAL_EXCEL_PATH})")

        processed_count += 1
        time.sleep(1)  # Rate limiting pause

if __name__ == "__main__":
    run_agent(limit=5)