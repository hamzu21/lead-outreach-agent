import os
import json
import time
import datetime
from src.config import (
    EXPENSE_SPREADSHEET_ID,
    EXPENSE_SHEET_NAME,
    EXPENSE_EXCEL_PATH,
    LOCAL_EXCEL_PATH
)
from src.services.google_auth import get_google_services
from src.services.gmail_service import send_gmail_message, create_gmail_draft
from src.services.workspace_service import (
    create_google_doc,
    update_google_doc,
    create_styled_spreadsheet,
    update_spreadsheet_data,
    trash_drive_file,
    list_workspace_files
)
from src.services.ai_generator import generate_ai_content
from src.services.telegram_service import send_telegram_message
from openpyxl import Workbook, load_workbook

def extract_clean_email(text: str) -> str:
    """Extracts raw email address from header string like 'Duolingo <hello@duolingo.com>'"""
    import re
    match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', text)
    if match:
        return match.group(0)
    return text.strip()


class PersonalAssistantService:
    def __init__(self):
        self.sheets_service = None
        self.gmail_service = None
        self.docs_service = None
        self.drive_service = None

    def initialize_services(self):
        try:
            self.sheets_service, self.gmail_service, self.docs_service, self.drive_service = get_google_services()
        except Exception as e:
            print(f"[PersonalAssistant] Warning initializing Google services: {e}")

    def fetch_recent_emails(self, max_results: int = 15, query: str = "label:INBOX"):
        """
        Fetches metadata of recent emails from Gmail.
        """
        if not self.gmail_service:
            self.initialize_services()

        if not self.gmail_service:
            return []

        try:
            res = self.gmail_service.users().messages().list(
                userId="me",
                q=query,
                maxResults=max_results
            ).execute()
            messages = res.get("messages", [])

            # Fallback search if query label:INBOX returns empty
            if not messages and query == "label:INBOX":
                res = self.gmail_service.users().messages().list(
                    userId="me",
                    q="in:inbox",
                    maxResults=max_results
                ).execute()
                messages = res.get("messages", [])

            if not messages:
                res = self.gmail_service.users().messages().list(
                    userId="me",
                    q="",
                    maxResults=max_results
                ).execute()
                messages = res.get("messages", [])

            email_data = []
            for msg_meta in messages:
                msg = self.gmail_service.users().messages().get(
                    userId="me",
                    id=msg_meta["id"],
                    format="full"
                ).execute()

                payload = msg.get("payload", {})
                headers = payload.get("headers", [])

                subject = ""
                sender = ""
                date_str = ""

                for h in headers:
                    h_name = h.get("name", "").lower()
                    if h_name == "subject":
                        subject = h.get("value", "")
                    elif h_name == "from":
                        sender = h.get("value", "")
                    elif h_name == "date":
                        date_str = h.get("value", "")

                snippet = msg.get("snippet", "")
                email_data.append({
                    "id": msg_meta["id"],
                    "subject": subject,
                    "sender": sender,
                    "date": date_str,
                    "snippet": snippet
                })
            return email_data
        except Exception as e:
            print(f"[PersonalAssistant] Error fetching emails: {e}")
            return []

    def get_morning_briefing(self) -> str:
        """
        Generates a comprehensive Morning Executive Briefing.
        """
        print("[PersonalAssistant] Generating Morning Briefing...")
        emails = self.fetch_recent_emails(max_results=10, query="is:unread label:INBOX")

        now_str = datetime.datetime.now().strftime("%A, %B %d, %Y")
        
        prompt = f"""
You are an executive AI personal assistant. Create a clear, high-priority Morning Executive Briefing for today ({now_str}).

Context Data:
Unread Inbox Emails: {json.dumps(emails, indent=2)}

Requirements:
1. Provide a warm, professional morning greeting.
2. Highlight High Priority items / actionable emails (if any).
3. Provide a short summary of General Updates.
4. Give a motivating Tech/Coding Quote or Quick Insight for the day.
5. Format the output neatly using Telegram Markdown (use *bold*, bullet points, short paragraphs).

Keep it concise and easy to read on a mobile phone.
"""
        briefing_text = generate_ai_content(prompt)
        return briefing_text

    def process_expense(self, input_text: str) -> dict:
        """
        Extracts structured expense details from text/receipt using Gemini AI,
        and logs it to Google Sheet & Local Excel.
        """
        print(f"[PersonalAssistant] Processing expense input: {input_text[:50]}...")
        now_str = datetime.datetime.now().strftime("%Y-%m-%d")

        prompt = f"""
Extract structured financial transaction data from the following text or receipt details:
"{input_text}"

Current Date: {now_str}

Extract and return JSON with keys:
- "date": "YYYY-MM-DD" (use current date if unspecified)
- "vendor": "Name of store/vendor/payee"
- "amount": numeric float value
- "currency": "PKR", "USD", etc.
- "category": Choose one of ["Food & Dining", "Tech & Subscriptions", "Utilities & Bills", "Travel & Transport", "Education & Books", "Personal & Misc"]
- "description": Short 1-sentence description

Return strict JSON:
{{
  "date": "2026-08-22",
  "vendor": "Vendor Name",
  "amount": 50.0,
  "currency": "USD",
  "category": "Category",
  "description": "Short description"
}}
"""
        raw_json = generate_ai_content(prompt, response_mime_type="application/json")
        data = json.loads(raw_json)

        # Log to Google Sheet & Local Excel
        self._log_expense_data(data)
        return data

    def _log_expense_data(self, data: dict):
        """
        Logs extracted expense row to Google Sheet and local expense_log.xlsx
        """
        date_val = data.get("date", "")
        vendor_val = data.get("vendor", "Unknown")
        amount_val = data.get("amount", 0.0)
        currency_val = data.get("currency", "USD")
        category_val = data.get("category", "Personal")
        desc_val = data.get("description", "")

        row = [date_val, vendor_val, amount_val, currency_val, category_val, desc_val]

        # 1. Update Google Sheet
        if not self.sheets_service:
            self.initialize_services()

        if self.sheets_service:
            try:
                # Ensure sheet tab exists
                body = {"values": [row]}
                self.sheets_service.spreadsheets().values().append(
                    spreadsheetId=EXPENSE_SPREADSHEET_ID,
                    range=f"{EXPENSE_SHEET_NAME}!A:F",
                    valueInputOption="USER_ENTERED",
                    insertDataOption="INSERT_ROWS",
                    body=body
                ).execute()
                print("-> Appended expense to Google Sheet.")
            except Exception as e:
                print(f"Warning: Failed to update expense in Google Sheet: {e}")

        # 2. Update Local Excel
        try:
            excel_path = EXPENSE_EXCEL_PATH
            if not os.path.exists(excel_path):
                wb = Workbook()
                ws = wb.active
                ws.title = "Expenses"
                ws.append(["Date", "Vendor", "Amount", "Currency", "Category", "Description"])
            else:
                wb = load_workbook(excel_path)
                ws = wb.active

            ws.append(row)
            wb.save(excel_path)
            print(f"-> Appended expense to local Excel ({excel_path}).")
        except Exception as e:
            print(f"Warning: Failed to update local expense Excel: {e}")

    def fetch_drafts(self, max_results: int = 15) -> list:
        """
        Fetches metadata of saved email drafts from Gmail.
        """
        if not self.gmail_service:
            self.initialize_services()

        if not self.gmail_service:
            return []

        try:
            res = self.gmail_service.users().drafts().list(
                userId="me",
                maxResults=max_results
            ).execute()
            drafts = res.get("drafts", [])

            draft_data = []
            for d in drafts:
                draft_id = d["id"]
                draft_obj = self.gmail_service.users().drafts().get(
                    userId="me",
                    id=draft_id
                ).execute()

                msg = draft_obj.get("message", {})
                msg_id = msg.get("id", "")
                payload = msg.get("payload", {})
                headers = payload.get("headers", [])

                subject = ""
                to_email = ""
                date_str = ""

                for h in headers:
                    h_name = h.get("name", "").lower()
                    if h_name == "subject":
                        subject = h.get("value", "")
                    elif h_name == "to":
                        to_email = h.get("value", "")
                    elif h_name == "date":
                        date_str = h.get("value", "")

                snippet = msg.get("snippet", "")
                draft_data.append({
                    "draft_id": draft_id,
                    "msg_id": msg_id,
                    "subject": subject or "(No Subject)",
                    "to": to_email or "Unspecified",
                    "date": date_str,
                    "snippet": snippet
                })
            return draft_data
        except Exception as e:
            print(f"[PersonalAssistant] Error fetching drafts: {e}")
            return []

    def get_drafts_digest(self) -> str:
        """
        Formats a clean Telegram digest of saved Gmail drafts.
        """
        print("[PersonalAssistant] Generating Drafts Digest...")
        drafts = self.fetch_drafts(max_results=15)

        if not drafts:
            return "📝 *Email Drafts*: No saved drafts found in your Gmail account."

        res_text = f"📝 *Saved Email Drafts ({len(drafts)} Found)*:\n\n"
        for idx, d in enumerate(drafts, start=1):
            res_text += (
                f"*{idx}. {d['subject']}*\n"
                f"• *To*: `{d['to']}`\n"
                f"• *Snippet*: {d['snippet'][:100]}...\n"
                f"• *Draft ID*: `{d['draft_id']}`\n\n"
            )
        res_text += "💡 *Tip*: Tell me to send any draft, e.g., _'Send draft 1'_ or _'Send draft r12345'_."
        return res_text

    def send_draft(self, draft_id: str) -> bool:
        """
        Sends an existing Gmail draft by draft_id.
        """
        if not self.gmail_service:
            self.initialize_services()
        if not self.gmail_service:
            return False

        try:
            self.gmail_service.users().drafts().send(
                userId="me",
                body={"id": draft_id}
            ).execute()
            print(f"-> Draft {draft_id} sent successfully!")
            return True
        except Exception as e:
            print(f"[PersonalAssistant] Error sending draft {draft_id}: {e}")
            return False

    def trash_email(self, message_id: str) -> bool:
        """
        Moves a message/draft to Trash.
        """
        if not self.gmail_service:
            self.initialize_services()
        if not self.gmail_service:
            return False

    def send_email(self, to_email: str, subject: str, body_text: str) -> dict:
        """
        Composes and sends an email directly via Gmail API.
        """
        if not self.gmail_service:
            self.initialize_services()
        if not self.gmail_service:
            return {"success": False, "error": "Gmail service unavailable"}

        try:
            msg_id = send_gmail_message(self.gmail_service, to_email, subject, body_text)
            print(f"-> Sent email to {to_email} (Msg ID: {msg_id})")
            return {"success": True, "msg_id": msg_id}
        except Exception as e:
            print(f"[PersonalAssistant] Error sending email to {to_email}: {e}")
            return {"success": False, "error": str(e)}

    def resolve_target_email(self, keyword_or_email: str) -> dict:
        """
        Smart recipient lookup. If given a full email address (e.g. 'zeusmr777@gmail.com'), returns it.
        If given a brand/name/keyword (e.g. 'duolingo', 'linkedin', 'rocketams'), queries Gmail
        to find the most recent matching email and extracts the recipient/sender email address,
        automatically skipping the account owner's own email address.
        """
        cleaned_kw = (keyword_or_email or "").strip(' "\'\t\r\n')
        
        # 1. If already a valid email address
        if "@" in cleaned_kw and "." in cleaned_kw.split("@")[-1]:
            raw_email = extract_clean_email(cleaned_kw)
            return {"to_email": raw_email, "subject": "Inquiry"}

        # 2. Search Gmail for matching messages by keyword/brand name
        search_q = cleaned_kw if cleaned_kw else "label:INBOX"
        try:
            if not self.gmail_service:
                self.initialize_services()
            if self.gmail_service:
                # Fetch user's own email address to prevent sending email to self
                my_email = ""
                try:
                    prof = self.gmail_service.users().getProfile(userId="me").execute()
                    my_email = prof.get("emailAddress", "").lower().strip()
                except Exception:
                    pass

                res = self.gmail_service.users().messages().list(
                    userId="me",
                    q=search_q,
                    maxResults=10
                ).execute()
                messages = res.get("messages", [])
                
                for msg_meta in messages:
                    msg = self.gmail_service.users().messages().get(
                        userId="me",
                        id=msg_meta["id"],
                        format="full"
                    ).execute()
                    
                    headers = msg.get("payload", {}).get("headers", [])
                    sender_email = ""
                    recipient_email = ""
                    subject_str = ""
                    for h in headers:
                        h_name = h.get("name", "").lower()
                        if h_name == "from":
                            sender_email = extract_clean_email(h.get("value", ""))
                        elif h_name == "to":
                            recipient_email = extract_clean_email(h.get("value", ""))
                        elif h_name == "subject":
                            subject_str = h.get("value", "")
                            
                    # Target the email address that is NOT the account owner
                    target_email = ""
                    if sender_email and sender_email.lower() != my_email:
                        target_email = sender_email
                    elif recipient_email and recipient_email.lower() != my_email:
                        target_email = recipient_email

                    if target_email and "@" in target_email and target_email.lower() != my_email:
                        try:
                            print(f"-> Smart resolved '{cleaned_kw}' to target '{target_email}'")
                        except Exception:
                            pass
                        return {"to_email": target_email, "subject": subject_str}
        except Exception as e:
            print(f"[PersonalAssistant] Error resolving target email for '{cleaned_kw}': {e}")

        return {"to_email": cleaned_kw, "subject": "Inquiry"}

    def reply_to_email(self, to_email: str, subject: str, instructions: str) -> dict:
        """
        Generates a smart AI response using Gemini based on user instructions,
        and sends it directly to the target recipient email via Gmail API.
        """
        print(f"[PersonalAssistant] Generating AI reply for {to_email}...")
        
        reply_subject = subject if (subject and subject.startswith("Re:")) else f"Re: {subject or 'Your Inquiry'}"
        
        prompt = f"""
You are an executive assistant composing a professional email response on behalf of Muhammad Hamza.

Recipient Email: {to_email}
Original Subject: {subject}
User's Instructions for Reply: "{instructions}"

Write a polished, professional, clear, and friendly email body text. Do not include markdown code block quotes, just the raw email text.
"""
        generated_body = generate_ai_content(prompt)
        
        # Send email via Gmail service
        res = self.send_email(to_email=to_email, subject=reply_subject, body_text=generated_body)
        if res.get("success"):
            res["generated_body"] = generated_body
            res["subject"] = reply_subject
        return res

    def get_inbox_digest(self) -> str:
        """
        Categorizes inbox emails and builds a structured summary.
        """
        print("[PersonalAssistant] Generating Inbox Digest...")
        emails = self.fetch_recent_emails(max_results=15, query="label:INBOX")

        if not emails:
            return "📥 *Inbox Digest*: No recent inbox messages found."

        prompt = f"""
Categorize and summarize the following unread/recent inbox emails:
{json.dumps(emails, indent=2)}

Group emails into the following categories:
- 🚨 *Action Required* (Emails needing direct reply or urgent action)
- 💳 *Bills & Financial* (Invoices, receipts, subscription notices)
- 👤 *Personal / Important* (Direct contacts)
- 📰 *Newsletters / Info* (General updates, non-urgent)

Format the response neatly for Telegram (Markdown). Keep summaries punchy and clear.
"""
        digest_text = generate_ai_content(prompt)
        return digest_text

    def create_google_document(self, title: str, instructions: str) -> dict:
        """
        Uses Gemini to generate well-structured document text, creates a new Google Doc,
        makes it shareable, and returns the view URL.
        """
        if not self.docs_service or not self.drive_service:
            self.initialize_services()

        print(f"[PersonalAssistant] Drafting Google Doc: '{title}'...")
        prompt = f"""
You are an executive assistant drafting a comprehensive, well-structured document on behalf of Muhammad Hamza.

Document Title: {title}
Topic / User Instructions: "{instructions}"

Write a detailed, clear, professional document.
Use clean section headers (e.g., # Executive Summary, ## Section 1), bullet points, and short readable paragraphs.
Do not wrap in markdown code blocks, just return raw document text.
"""
        doc_text = generate_ai_content(prompt)
        return create_google_doc(self.docs_service, self.drive_service, title=title, content_text=doc_text)

    def create_google_sheet(self, title: str, instructions: str) -> dict:
        """
        Uses Gemini to generate structured tabular data (headers + rows),
        creates a styled Google Sheet with colored headers, makes it shareable,
        and returns the spreadsheet URL.
        """
        if not self.sheets_service or not self.drive_service:
            self.initialize_services()

        print(f"[PersonalAssistant] Generating Google Sheet: '{title}'...")
        prompt = f"""
You are an executive data assistant creating a structured spreadsheet table for Muhammad Hamza.

Spreadsheet Title: {title}
Data Description / User Instructions: "{instructions}"

Generate structured JSON containing:
- "headers": List of string column header names (e.g. ["Category", "Item", "Cost (USD)", "Status"])
- "rows": List of rows, where each row is a list of cell values
- "theme_color": Choose one of ["blue", "green", "purple"]

Return strict JSON:
{{
  "headers": ["Header 1", "Header 2", "Header 3"],
  "rows": [
    ["Val 1", "Val 2", "Val 3"],
    ["Val 4", "Val 5", "Val 6"]
  ],
  "theme_color": "blue"
}}
"""
        raw_json = generate_ai_content(prompt, response_mime_type="application/json")
        try:
            clean_str = raw_json.strip().strip("`").replace("json\n", "")
            data = json.loads(clean_str)
            headers = data.get("headers", ["Item", "Details", "Status"])
            rows = data.get("rows", [])
            theme = data.get("theme_color", "blue")
        except Exception as e:
            print(f"[PersonalAssistant] Error parsing Gemini sheet JSON: {e}")
            headers = ["Title", "Details", "Date"]
            rows = [[title, instructions, datetime.datetime.now().strftime("%Y-%m-%d")]]
            theme = "blue"

        return create_styled_spreadsheet(
            self.sheets_service,
            self.drive_service,
            title=title,
            headers=headers,
            rows=rows,
            theme_color=theme
        )

    def trash_workspace_file(self, file_identifier: str) -> dict:
        """
        Moves a Google Doc or Google Sheet file to Drive Trash by ID or title keyword.
        """
        if not self.drive_service:
            self.initialize_services()
        if not self.drive_service:
            return {
                "success": False,
                "error": "Google Drive permissions needed. Please run 'python reauth_google.py' once on your PC to authorize Drive access."
            }
        return trash_drive_file(self.drive_service, file_identifier)

    def list_workspace_files_digest(self, file_type: str = "spreadsheet") -> str:
        """
        Lists Google Drive files (spreadsheets or documents) and formats a clean Telegram digest with URLs.
        """
        if not self.drive_service:
            self.initialize_services()
        
        res = list_workspace_files(self.drive_service, file_type=file_type)
        if not res.get("success"):
            return f"⚠️ Could not list Drive files: {res.get('error')}"

        files = res.get("files", [])
        if not files:
            return f"📂 *Google Drive*: No active {file_type} files found in your cloud drive."

        type_label = "Spreadsheets" if "sheet" in file_type.lower() else "Workspace Files"
        res_text = f"📊 *Your Google Cloud {type_label} ({len(files)} Found)*:\n\n"
        for idx, f in enumerate(files, start=1):
            name = f.get("name", "Untitled")
            f_id = f.get("id")
            link = f.get("webViewLink") or f"https://docs.google.com/spreadsheets/d/{f_id}/edit"
            res_text += f"*{idx}. {name}*\n• 🔗 [Open File]({link})\n• *ID*: `{f_id}`\n\n"

        return res_text


def run_morning_brief_agent(send_telegram: bool = True):
    service = PersonalAssistantService()
    brief = service.get_morning_briefing()
    print("\n--- MORNING EXECUTIVE BRIEFING ---")
    print(brief)
    if send_telegram:
        success = send_telegram_message(brief)
        if success:
            print("[PersonalAssistant] Morning Briefing sent to Telegram successfully.")

def run_expense_tracker_agent(input_text: str, send_telegram: bool = True) -> dict:
    service = PersonalAssistantService()
    data = service.process_expense(input_text)
    msg = (
        f"✅ *Expense Logged Successfully*\n\n"
        f"• *Vendor*: {data.get('vendor')}\n"
        f"• *Amount*: {data.get('currency')} {data.get('amount')}\n"
        f"• *Category*: {data.get('category')}\n"
        f"• *Date*: {data.get('date')}\n"
        f"• *Note*: {data.get('description')}"
    )
    print("\n--- EXPENSE PROCESSED ---")
    print(msg)
    if send_telegram:
        send_telegram_message(msg)
    return data

def run_inbox_zero_agent(send_telegram: bool = True):
    service = PersonalAssistantService()
    digest = service.get_inbox_digest()
    print("\n--- INBOX DIGEST ---")
    print(digest)
    if send_telegram:
        send_telegram_message(digest)
