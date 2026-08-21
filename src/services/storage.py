import os
import datetime
from openpyxl import Workbook, load_workbook
from src.config import SPREADSHEET_ID, SHEET_NAME, LOCAL_EXCEL_PATH

def ensure_google_sheet_header(sheets_service):
    """
    Ensures Column I ('Outreach Status') and Column J ('WhatsApp Link') headers exist in the Google Sheet.
    """
    try:
        header_res = sheets_service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_ID,
            range=f"{SHEET_NAME}!I1:J1"
        ).execute()
        vals = header_res.get("values", [])
        if not vals or len(vals[0]) < 2:
            sheets_service.spreadsheets().values().update(
                spreadsheetId=SPREADSHEET_ID,
                range=f"{SHEET_NAME}!I1:J1",
                valueInputOption="RAW",
                body={"values": [["Outreach Status", "WhatsApp Link"]]}
            ).execute()
    except Exception as e:
        print(f"Warning: Could not check/set headers: {e}")

def update_google_sheet_status(sheets_service, row_idx: int, draft_id: str, whatsapp_link: str = "N/A"):
    """
    Updates Column I (Outreach Status) and Column J (WhatsApp Link) in Google Sheet.
    """
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    update_range = f"{SHEET_NAME}!I{row_idx}:J{row_idx}"
    sheets_service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=update_range,
        valueInputOption="RAW",
        body={"values": [[f"Draft Created [{now_str}] (ID: {draft_id})", whatsapp_link]]}
    ).execute()

def update_local_excel(business_name: str, email: str, location: str, industry: str, draft_id: str, subject: str, whatsapp_link: str = "N/A"):
    """
    Appends or updates a lead's record in the local Excel spreadsheet log file.
    """
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    headers = ["Business Name", "Email Address", "Location", "Industry", "Draft ID", "Subject Line", "Status", "Timestamp", "WhatsApp Link"]
    
    if os.path.exists(LOCAL_EXCEL_PATH):
        try:
            wb = load_workbook(LOCAL_EXCEL_PATH)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Drafted Leads"
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Drafted Leads"
        ws.append(headers)

    # Check if entry already exists to update
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if len(row) > 1 and row[1].value == email:
            row[4].value = draft_id
            row[5].value = subject
            row[6].value = "Draft Created"
            row[7].value = now_str
            if len(row) > 8:
                row[8].value = whatsapp_link
            else:
                ws.cell(row=row[0].row, column=9, value=whatsapp_link)
            found = True
            break

    if not found:
        ws.append([business_name, email, location, industry, draft_id, subject, "Draft Created", now_str, whatsapp_link])

    wb.save(LOCAL_EXCEL_PATH)
